from openpyxl import load_workbook
from collections import Counter

def check_row_conditions(row):
    numbers = [int(cell.value) for cell in row]

    num_counts = Counter(numbers)

    max_num = max(numbers)

    has_repeated_numbers = any(count > 1 for count in num_counts.values())

    max_not_repeated = num_counts[max_num] == 1

    sum_repeated = sum(num * count for num, count in num_counts.items() if count > 1)

    return has_repeated_numbers and max_not_repeated and sum_repeated < max_num

def count_valid_rows(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active

    valid_row_count = 0

    for row in sheet.iter_rows(min_row=1, max_col=6):
        if check_row_conditions(row):
            valid_row_count += 1

    return valid_row_count

file_path = '09.xlsx'
result = count_valid_rows(file_path)
print(result)
