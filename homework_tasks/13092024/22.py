import openpyxl

df = openpyxl.load_workbook("22.xlsx")["Лист1"]

instant_process = {}

for i in range(2, df.max_row + 1):
    idd = df[f"A{i}"]
    sender_id = df[f"C{i}"]
    time = df[f"B{i}"]

    if sender_id.value == 0:
        instant_process[idd.value] = time.value

print(instant_process.keys())

counter = 0

for i in range(2, df.max_row + 1):
    idd = df[f"A{i}"]
    sender_id = df[f"C{i}"]
    time = df[f"B{i}"]

    if sender_id.value == 0:
        pass
    else:
        time_required = 0
        sender_id = str(sender_id.value).split(";")
        print(sender_id)
        for j in range(len(sender_id)):
            time_required += instant_process[int(sender_id[j])]
        time_required += time.value
        if time_required >= 100:
            counter += 1
        print("id:", idd.value, "time:", time_required)
        instant_process[idd.value] = time_required
print(counter)