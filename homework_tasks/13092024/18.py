import openpyxl

df = openpyxl.load_workbook("18.xlsx")["18"]

def where_to_step_max(x, y, summ):
    curr_value = df.cell(row=y, column=x).value

    if x == df.max_column and y == df.max_row:
        return summ

    if x < df.max_column:
        right_value = df.cell(row=y, column=x + 1).value
    else:
        right_value = -float('inf')

    if y < df.max_row:
        down_value = df.cell(row=y + 1, column=x).value
    else:
        down_value = -float('inf')

    if right_value > down_value:
        if right_value > curr_value:
            summ += right_value
        else:
            summ -= right_value
        return where_to_step_max(x + 1, y, summ)
    else:
        if down_value > curr_value:
            summ += down_value
        else:
            summ -= down_value
        return where_to_step_max(x, y + 1, summ)

def where_to_step_min(x, y, summ):
    curr_value = df.cell(row=y, column=x).value

    if x == df.max_column and y == df.max_row:
        return summ

    if x < df.max_column:
        right_value = df.cell(row=y, column=x + 1).value
    else:
        right_value = float('inf')

    if y < df.max_row:
        down_value = df.cell(row=y + 1, column=x).value
    else:
        down_value = float('inf')

    if right_value < down_value:
        if right_value > curr_value:
            summ += right_value
        else:
            summ -= right_value
        return where_to_step_min(x + 1, y, summ)
    else:
        if down_value > curr_value:
            summ += down_value
        else:
            summ -= down_value
        return where_to_step_min(x, y + 1, summ)

x, y = 1, 1
start_value = df.cell(row=y, column=x).value

max_energy = where_to_step_max(x, y, start_value)
min_energy = where_to_step_min(x, y, start_value)

print(max_energy)
print(min_energy)
